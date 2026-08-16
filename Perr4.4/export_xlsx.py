"""
Экспортирует общую таблицу заявок requests.csv в requests.xlsx — удобнее
открывать и показывать в Excel/Google Таблицах, чем сырой CSV.

Запуск (из папки Perr4.4): python export_xlsx.py
"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CSV_PATH = Path(__file__).resolve().parent / "requests.csv"
XLSX_PATH = Path(__file__).resolve().parent / "requests.xlsx"


def main() -> None:
    with open(CSV_PATH, newline="", encoding="utf-8-sig") as f:
        rows = [row for row in csv.reader(f) if row]

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F6FE0")

    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    widths = [6, 18, 20, 18, 48, 18, 18]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(XLSX_PATH)
    print(f"Сохранено: {XLSX_PATH} ({len(rows) - 1} заявок)")


if __name__ == "__main__":
    main()
